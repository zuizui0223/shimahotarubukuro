#!/usr/bin/env python3
"""Attach field coordinates from an Excel workbook to image-derived result tables.

The workbook may use Japanese or English column names. Coordinates are normalized
once, then joined conservatively to result rows using the strongest available key:

1. island + sheet + plant_id
2. island + plant_id
3. island + sheet
4. island only (only when the workbook has one unique coordinate for the island)

Ambiguous keys are never guessed. They are written to location_join_report.csv.
The original result files are not modified; coordinate-enriched copies are created.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from statistics import fmean
from typing import Any, Iterable

from openpyxl import load_workbook


ISLAND_ALIASES = {
    "oshima": "Oshima",
    "izuoshima": "Oshima",
    "伊豆大島": "Oshima",
    "大島": "Oshima",
    "toshima": "Toshima",
    "利島": "Toshima",
    "niijima": "Niijima",
    "新島": "Niijima",
    "shikinejima": "Shikinejima",
    "式根島": "Shikinejima",
    "kozushima": "Kozushima",
    "kouzushima": "Kozushima",
    "神津島": "Kozushima",
}

COLUMN_ALIASES = {
    "island": ["island", "island_name", "region", "島", "島名", "地域"],
    "sheet": [
        "sheet", "scan", "scan_id", "image", "image_id", "filename", "file_name",
        "file", "スキャン", "画像", "画像名", "ファイル", "ファイル名", "シート名",
    ],
    "plant_id": [
        "plant_id", "plant", "individual_id", "individual", "sample_id", "sample",
        "株id", "株番号", "株", "個体id", "個体番号", "個体", "サンプルid", "サンプル番号",
    ],
    "site_id": [
        "site_id", "site", "location_id", "location", "locality", "population", "pop",
        "地点id", "地点番号", "地点名", "地点", "採集地点", "調査地点", "個体群", "集団",
    ],
    "latitude": ["latitude", "lat", "緯度", "北緯"],
    "longitude": ["longitude", "lon", "lng", "long", "経度", "東経"],
    "coordinates": [
        "coordinates", "coordinate", "gps", "latlon", "lat_lon", "緯度経度", "座標", "gps座標",
    ],
    "elevation_m": ["elevation_m", "elevation", "altitude_m", "altitude", "標高m", "標高"],
    "coord_datum": ["coord_datum", "datum", "測地系", "座標系"],
    "coord_accuracy_m": [
        "coord_accuracy_m", "accuracy_m", "coordinate_accuracy", "gps_accuracy", "精度m", "gps精度",
    ],
    "coordinate_notes": ["coordinate_notes", "location_notes", "notes", "note", "備考", "メモ"],
}

OUTPUT_LOCATION_FIELDS = [
    "site_id", "latitude", "longitude", "elevation_m", "coord_datum",
    "coord_accuracy_m", "coordinate_notes", "coordinate_source",
    "coordinate_match_level",
]

NUMERIC_SUMMARY_FIELDS = [
    "corolla_len_mm", "corolla_width_mm", "corolla_area_mm2", "guide_area_mm2",
    "guide_cov_pct", "n_spots", "spot_density_cm2", "guide_extent_rel",
    "guide_present", "guide_cov_incl_oxidized_pct", "brown_frac", "degraded_flag",
]


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def normalized_token(value: Any) -> str:
    s = unicodedata.normalize("NFKC", text(value)).strip().lower()
    return re.sub(r"[^0-9a-zA-Z一-龥ぁ-んァ-ヶ]+", "", s)


def normalized_header(value: Any) -> str:
    return normalized_token(value)


ALIAS_LOOKUP: dict[str, str] = {}
for canonical, aliases in COLUMN_ALIASES.items():
    for alias in aliases:
        ALIAS_LOOKUP[normalized_header(alias)] = canonical


def canonical_island(value: Any) -> str:
    raw = unicodedata.normalize("NFKC", text(value)).strip()
    if not raw:
        return ""
    key = normalized_token(raw)
    return ISLAND_ALIASES.get(key, raw)


def normalize_sheet(value: Any) -> str:
    raw = unicodedata.normalize("NFKC", text(value)).strip()
    if not raw:
        return ""
    raw = Path(raw).stem
    return re.sub(r"[^0-9a-z]+", "", raw.lower())


def normalize_plant(value: Any) -> str:
    raw = unicodedata.normalize("NFKC", text(value)).strip()
    if not raw:
        return ""
    if re.fullmatch(r"[-+]?\d+\.0+", raw):
        raw = raw.split(".", 1)[0]
    return re.sub(r"\s+", "", raw)


def to_float(value: Any) -> float | None:
    if value is None or text(value) == "":
        return None
    if isinstance(value, (int, float)):
        out = float(value)
        return out if math.isfinite(out) else None
    s = unicodedata.normalize("NFKC", text(value)).replace(",", "")
    match = re.search(r"[-+]?\d+(?:\.\d+)?", s)
    if not match:
        return None
    try:
        out = float(match.group())
    except ValueError:
        return None
    return out if math.isfinite(out) else None


def parse_coord(value: Any, axis: str) -> float | None:
    """Parse decimal or DMS latitude/longitude, respecting hemisphere letters."""
    if value is None or text(value) == "":
        return None
    if isinstance(value, (int, float)):
        out = float(value)
        return out if math.isfinite(out) else None

    s = unicodedata.normalize("NFKC", text(value)).upper().strip()
    sign = -1.0 if any(h in s for h in ("S", "W", "南緯", "西経")) else 1.0
    nums = [float(x) for x in re.findall(r"[-+]?\d+(?:\.\d+)?", s)]
    if not nums:
        return None

    if len(nums) >= 2 and any(mark in s for mark in ("°", "度", "'", "′", '"', "″")):
        deg = abs(nums[0])
        minutes = abs(nums[1]) if len(nums) >= 2 else 0.0
        seconds = abs(nums[2]) if len(nums) >= 3 else 0.0
        out = deg + minutes / 60.0 + seconds / 3600.0
        if nums[0] < 0:
            sign = -1.0
        out *= sign
    else:
        out = nums[0]
        if any(h in s for h in ("S", "W", "南緯", "西経")):
            out = -abs(out)

    limit = 90.0 if axis == "latitude" else 180.0
    return out if -limit <= out <= limit else None


def parse_coordinate_pair(value: Any) -> tuple[float | None, float | None]:
    s = unicodedata.normalize("NFKC", text(value)).strip()
    if not s:
        return None, None

    parts = re.split(r"\s*[,，;/]\s*", s, maxsplit=1)
    if len(parts) == 2:
        lat = parse_coord(parts[0], "latitude")
        lon = parse_coord(parts[1], "longitude")
        if lat is not None and lon is not None:
            return lat, lon

    nums = [float(x) for x in re.findall(r"[-+]?\d+(?:\.\d+)?", s)]
    if len(nums) >= 2:
        lat, lon = nums[0], nums[1]
        if -90 <= lat <= 90 and -180 <= lon <= 180:
            return lat, lon
        if -90 <= lon <= 90 and -180 <= lat <= 180:
            return lon, lat
    return None, None


def map_headers(headers: Iterable[Any]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for idx, header in enumerate(headers):
        canonical = ALIAS_LOOKUP.get(normalized_header(header))
        if canonical and canonical not in mapping.values():
            mapping[idx] = canonical
    return mapping


def row_has_content(values: Iterable[Any]) -> bool:
    return any(text(v) for v in values)


def read_locations(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    records: list[dict[str, str]] = []
    warnings: list[str] = []

    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        header_values = None
        header_row_number = 0
        header_map: dict[int, str] = {}

        for row_number, values in enumerate(rows, start=1):
            if not row_has_content(values):
                continue
            candidate = map_headers(values)
            if ("latitude" in candidate.values() and "longitude" in candidate.values()) or \
                    "coordinates" in candidate.values():
                header_values = list(values)
                header_row_number = row_number
                header_map = candidate
                break
            if row_number >= 30:
                break

        if header_values is None:
            warnings.append(f"{ws.title}: no latitude/longitude or coordinates header found; skipped")
            continue

        raw_headers = [text(h) or f"column_{i + 1}" for i, h in enumerate(header_values)]
        for row_number, values in enumerate(rows, start=header_row_number + 1):
            if not row_has_content(values):
                continue
            values = list(values)
            canonical: dict[str, Any] = {
                canonical_name: values[idx] if idx < len(values) else None
                for idx, canonical_name in header_map.items()
            }

            lat = parse_coord(canonical.get("latitude"), "latitude")
            lon = parse_coord(canonical.get("longitude"), "longitude")
            if (lat is None or lon is None) and canonical.get("coordinates") is not None:
                pair_lat, pair_lon = parse_coordinate_pair(canonical.get("coordinates"))
                lat = lat if lat is not None else pair_lat
                lon = lon if lon is not None else pair_lon

            if lat is None or lon is None:
                warnings.append(f"{ws.title}!{row_number}: invalid or missing coordinate; skipped")
                continue

            extras = {}
            for idx, raw_header in enumerate(raw_headers):
                if idx in header_map or idx >= len(values):
                    continue
                if text(values[idx]):
                    extras[raw_header] = text(values[idx])

            elevation = to_float(canonical.get("elevation_m"))
            accuracy = to_float(canonical.get("coord_accuracy_m"))
            records.append({
                "island": canonical_island(canonical.get("island")),
                "sheet": text(canonical.get("sheet")),
                "sheet_key": normalize_sheet(canonical.get("sheet")),
                "plant_id": text(canonical.get("plant_id")),
                "plant_key": normalize_plant(canonical.get("plant_id")),
                "site_id": text(canonical.get("site_id")),
                "latitude": f"{lat:.7f}",
                "longitude": f"{lon:.7f}",
                "elevation_m": "" if elevation is None else f"{elevation:g}",
                "coord_datum": text(canonical.get("coord_datum")) or "WGS84",
                "coord_accuracy_m": "" if accuracy is None else f"{accuracy:g}",
                "coordinate_notes": text(canonical.get("coordinate_notes")),
                "coordinate_source": f"{path.name}:{ws.title}!{row_number}",
                "source_sheet": ws.title,
                "source_row": str(row_number),
                "extra_json": json.dumps(extras, ensure_ascii=False, sort_keys=True) if extras else "",
            })

    return records, warnings


def coordinate_signature(record: dict[str, str]) -> tuple[str, str, str, str]:
    return (
        record.get("latitude", ""), record.get("longitude", ""),
        record.get("site_id", ""), record.get("elevation_m", ""),
    )


def build_unique_index(
    records: list[dict[str, str]], key_fields: tuple[str, ...]
) -> tuple[dict[tuple[str, ...], dict[str, str]], set[tuple[str, ...]]]:
    groups: dict[tuple[str, ...], list[dict[str, str]]] = defaultdict(list)
    for record in records:
        key = tuple(record.get(field, "") for field in key_fields)
        if all(key):
            groups[key].append(record)

    unique: dict[tuple[str, ...], dict[str, str]] = {}
    ambiguous: set[tuple[str, ...]] = set()
    for key, rows in groups.items():
        signatures = {coordinate_signature(row) for row in rows}
        if len(signatures) == 1:
            unique[key] = rows[0]
        else:
            ambiguous.add(key)
    return unique, ambiguous


def target_identity(row: dict[str, str]) -> tuple[str, str, str]:
    island = canonical_island(row.get("island"))
    sheet = normalize_sheet(row.get("sheet"))
    plant = normalize_plant(row.get("plant_id") or row.get("plant_id_FILL"))
    return island, sheet, plant


def choose_location(
    row: dict[str, str],
    indexes: list[tuple[str, tuple[str, ...], dict[tuple[str, ...], dict[str, str]], set[tuple[str, ...]]]],
) -> tuple[dict[str, str] | None, str, str]:
    island, sheet, plant = target_identity(row)
    values = {"island": island, "sheet_key": sheet, "plant_key": plant}

    saw_ambiguous = []
    for level, fields, unique, ambiguous in indexes:
        key = tuple(values[field] for field in fields)
        if not all(key):
            continue
        if key in ambiguous:
            saw_ambiguous.append(level)
            continue
        if key in unique:
            return unique[key], level, ""

    if saw_ambiguous:
        return None, "", "ambiguous location rows at " + "; ".join(saw_ambiguous)
    missing = [name for name, val in (("island", island), ("sheet", sheet), ("plant_id", plant)) if not val]
    if missing:
        return None, "", "no unique match; target missing " + ", ".join(missing)
    return None, "", "no matching location row"


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), [dict(row) for row in reader]


def write_csv(path: Path, fieldnames: list[str], rows: Iterable[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def attach_to_dataset(
    source_path: Path,
    output_path: Path,
    indexes: list[tuple[str, tuple[str, ...], dict[tuple[str, ...], dict[str, str]], set[tuple[str, ...]]]],
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    fieldnames, rows = read_csv(source_path)
    output_fields = fieldnames + [field for field in OUTPUT_LOCATION_FIELDS if field not in fieldnames]
    joined_rows: list[dict[str, str]] = []
    report_rows: list[dict[str, str]] = []

    for row_number, row in enumerate(rows, start=2):
        location, match_level, reason = choose_location(row, indexes)
        out = dict(row)
        if location:
            for field in OUTPUT_LOCATION_FIELDS:
                if field == "coordinate_match_level":
                    out[field] = match_level
                else:
                    out[field] = location.get(field, "")
            status = "matched"
        else:
            for field in OUTPUT_LOCATION_FIELDS:
                out.setdefault(field, "")
            status = "unmatched"

        island, _, plant = target_identity(row)
        report_rows.append({
            "dataset": source_path.name,
            "row_number": str(row_number),
            "island": island,
            "sheet": row.get("sheet", ""),
            "plant_id": row.get("plant_id") or row.get("plant_id_FILL", ""),
            "status": status,
            "match_level": match_level,
            "reason": reason,
            "site_id": out.get("site_id", ""),
            "latitude": out.get("latitude", ""),
            "longitude": out.get("longitude", ""),
        })
        joined_rows.append(out)

    write_csv(output_path, output_fields, joined_rows)
    return joined_rows, report_rows


def summarize_by_location(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    groups: dict[tuple[str, str, str, str], list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        if not row.get("latitude") or not row.get("longitude"):
            continue
        key = (
            canonical_island(row.get("island")), row.get("site_id", ""),
            row.get("latitude", ""), row.get("longitude", ""),
        )
        groups[key].append(row)

    summaries: list[dict[str, str]] = []
    for (island, site_id, lat, lon), group in sorted(groups.items()):
        summary = {
            "island": island,
            "site_id": site_id,
            "latitude": lat,
            "longitude": lon,
            "n_corollas": str(len(group)),
            "n_plants": str(len({normalize_plant(r.get("plant_id")) for r in group if normalize_plant(r.get("plant_id"))})),
        }
        for field in NUMERIC_SUMMARY_FIELDS:
            values = [to_float(row.get(field)) for row in group]
            clean = [value for value in values if value is not None]
            summary[f"mean_{field}"] = "" if not clean else f"{fmean(clean):.6g}"
        summaries.append(summary)
    return summaries


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Attach coordinates from location.xlsx to shimahotarubukuro result CSVs."
    )
    parser.add_argument("--locations", type=Path, default=Path("location.xlsx"))
    parser.add_argument("--results-dir", type=Path, default=Path("results"))
    parser.add_argument(
        "--output-dir", type=Path, default=None,
        help="Default: RESULTS_DIR/with_coordinates",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    locations_path = args.locations.expanduser().resolve()
    results_dir = args.results_dir.expanduser().resolve()
    output_dir = (
        args.output_dir.expanduser().resolve()
        if args.output_dir is not None
        else results_dir / "with_coordinates"
    )

    if not locations_path.exists():
        raise SystemExit(f"Location workbook not found: {locations_path}")
    if not results_dir.exists():
        raise SystemExit(f"Results directory not found: {results_dir}")

    locations, warnings = read_locations(locations_path)
    if not locations:
        detail = "\n".join(warnings) if warnings else "No readable rows."
        raise SystemExit(f"No valid coordinate rows found in {locations_path}.\n{detail}")

    normalized_location_fields = [
        "island", "sheet", "plant_id", "site_id", "latitude", "longitude",
        "elevation_m", "coord_datum", "coord_accuracy_m", "coordinate_notes",
        "coordinate_source", "source_sheet", "source_row", "extra_json",
    ]
    write_csv(output_dir / "locations_normalized.csv", normalized_location_fields, locations)

    index_specs = [
        ("island+sheet+plant_id", ("island", "sheet_key", "plant_key")),
        ("island+plant_id", ("island", "plant_key")),
        ("island+sheet", ("island", "sheet_key")),
        ("island", ("island",)),
    ]
    indexes = []
    for level, fields in index_specs:
        unique, ambiguous = build_unique_index(locations, fields)
        indexes.append((level, fields, unique, ambiguous))

    datasets = ["traits.csv", "qc_plant_labels.csv", "styles.csv"]
    all_reports: list[dict[str, str]] = []
    joined_traits: list[dict[str, str]] = []
    processed = []
    for filename in datasets:
        source_path = results_dir / filename
        if not source_path.exists():
            continue
        output_path = output_dir / filename.replace(".csv", "_with_coordinates.csv")
        joined, report = attach_to_dataset(source_path, output_path, indexes)
        all_reports.extend(report)
        processed.append(filename)
        if filename == "traits.csv":
            joined_traits = joined

    if not processed:
        raise SystemExit(f"No result CSVs found in {results_dir}")

    report_fields = [
        "dataset", "row_number", "island", "sheet", "plant_id", "status",
        "match_level", "reason", "site_id", "latitude", "longitude",
    ]
    write_csv(output_dir / "location_join_report.csv", report_fields, all_reports)

    if joined_traits:
        summaries = summarize_by_location(joined_traits)
        summary_fields = [
            "island", "site_id", "latitude", "longitude", "n_corollas", "n_plants",
        ] + [f"mean_{field}" for field in NUMERIC_SUMMARY_FIELDS]
        write_csv(output_dir / "per_location_summary.csv", summary_fields, summaries)

    matched = sum(row["status"] == "matched" for row in all_reports)
    unmatched = len(all_reports) - matched
    print(f"Read {len(locations)} valid coordinate rows from {locations_path.name}")
    print(f"Processed: {', '.join(processed)}")
    print(f"Joined rows: {matched}; unmatched/ambiguous rows: {unmatched}")
    print(f"Outputs: {output_dir}")
    if warnings:
        print("Warnings:")
        for warning in warnings:
            print(f"  - {warning}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

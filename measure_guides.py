# -*- coding: utf-8 -*-
"""
measure_guides.py — Nectar-guide & floral-size trait extraction from flat-bed
scans of *Campanula microdonta* (シマホタルブクロ) corollas, Izu islands.

Design
------
Each scan sheet holds several opened/flattened corollas of one or more plants,
plus a printed cm ruler. All scans are 300 DPI (verified from JPEG metadata and
cross-checked against the ruler), so the pixel scale is fixed:

    px/cm = 300 / 2.54 = 118.11      (1 px = 0.08467 mm ; 1 px^2 = 0.007168 mm^2)

Sampling unit: PLANT (株). 1–2 flowers (corollas) per plant. Corollas are
detected and measured individually; plant grouping is finalised by the user in
the QC template (handwritten circled numbers on the sheets are the ground truth).

Traits
------
Robust (validated; island means of corolla length reproduce Inoue's independent
common-garden values for Oshima/Toshima):
  * corolla_len_mm, corolla_width_mm, corolla_area_mm2   (size / "display")
  * guide_area_mm2, guide_cov_pct, n_spots, spot_density_cm2,
    guide_extent_rel, guide_present                       (nectar guide)
  * brown_frac, degraded_flag                             (preservation control)

Provisional (UNRELIABLE from flattened, inconsistently folded specimens — torn
/curled edges defeat lobe & sinus detection; folded-vs-open changes width→girth.
Use fresh-flower calipers / lateral calibrated photos for these instead):
  * prov_mouth_diam_mm, prov_tube_depth_mm, prov_n_lobes

Colour is used ONLY to locate purple pigment (pigment index = a* - b* in CIELAB,
which cleanly separates magenta guide from orange-brown degradation and cream
tissue). No colour value is reported.

Reproductive organs (style/pistil sticks laid beside corollas) are detected as
thin elongated objects and written to a separate CSV; they REQUIRE manual QC
(link to plant, exclude tape/filament artefacts). Stamen length and herkogamy
are NOT recoverable from these detached, flattened specimens.

Usage
-----
    python measure_guides.py --data-root "path/to/shimahotarubukuro" \
                             --out-dir results

Data root layout (island sub-folders, English names):
    oshima/  toshima/  niijima/  shikinejima/  kozushima/   with *.jpg sheets.
"""
from __future__ import annotations
import argparse, os, math, csv, re, unicodedata
import numpy as np, cv2
from PIL import Image, ImageOps

# ---- fixed scale (300 DPI) ----
PXCM = 300 / 2.54
MM_PX = 10.0 / PXCM
MM2_PX = MM_PX ** 2

# ---- corolla acceptance filters ----
AREA_MM2_MIN, AREA_MM2_MAX = 80.0, 3200.0
ASPECT_MAX, SOLIDITY_MIN = 4.0, 0.45

# ---- island metadata (Izu isolation gradient; region_order after Inoue) ----
ISLANDS = {
    "oshima":     ("Oshima", 1),
    "toshima":    ("Toshima", 2),
    "niijima":    ("Niijima", 3),
    "shikinejima": ("Shikinejima", 3.5),   # not in Inoue table; geographically Niijima–Kozushima
    "kozushima":  ("Kozushima", 4),
}


# Canonical orientation: every sheet is normalised so the ruler sits at the TOP
# (ruler bottom -> 180 deg; ruler on the left -> 90 deg clockwise). Ruler position
# is fixed per scan, so it is recorded here by file stem rather than detected —
# specimen edges make automatic top/bottom detection unreliable. Sheets already
# ruler-at-top are absent (no rotation). Keyed by the lower-cased filename stem.
SHEET_ROTATION = {
    "oshima1":     cv2.ROTATE_180,
    "oshima1(2)":  cv2.ROTATE_180,
    "oshima4":     cv2.ROTATE_180,
    "oshima5":     cv2.ROTATE_180,
    "oshima7":     cv2.ROTATE_180,
    "oshima8~9":   cv2.ROTATE_180,
    "oshima13~15": cv2.ROTATE_180,
    "toshima3~6":  cv2.ROTATE_180,
    "oshima6":     cv2.ROTATE_90_CLOCKWISE,   # ruler on the left edge
    "oshima8":     cv2.ROTATE_90_CLOCKWISE,   # ruler on the left edge
    "oshima10~13": cv2.ROTATE_90_CLOCKWISE,   # ruler on the left edge (landscape)
}


def canonical_rotation(path):
    """cv2 rotate code that brings this sheet's ruler to the top, or None."""
    stem = os.path.splitext(os.path.basename(str(path)))[0].lower()
    return SHEET_ROTATION.get(stem)


def load_bgr(path):
    """Load a JPEG with its EXIF orientation applied, then rotate to the canonical
    ruler-at-top orientation so every sheet is processed the same way up."""
    im = ImageOps.exif_transpose(Image.open(path)).convert("RGB")
    img = cv2.cvtColor(np.array(im), cv2.COLOR_RGB2BGR)
    rotate = canonical_rotation(path)
    if rotate is not None:
        img = cv2.rotate(img, rotate)
    return img


def _lstd(gray, k=15):
    g = gray.astype(np.float32)
    m = cv2.boxFilter(g, -1, (k, k)); sq = cv2.boxFilter(g * g, -1, (k, k))
    return np.sqrt(np.clip(sq - m * m, 0, None))


def channels(img):
    lab = cv2.cvtColor(img, cv2.COLOR_BGR2LAB).astype(np.float32)
    return lab[..., 0], lab[..., 1] - 128, lab[..., 2] - 128     # L, a*, b*


def foreground(img):
    """Whole-corolla mask. Cream/translucent tissue is separated from white paper
    and the neutral metal ruler by chroma OR surface texture-with-a-little-colour."""
    Lc, a, b = channels(img)
    chroma = np.sqrt(a * a + b * b); tex = _lstd(Lc, 15)
    fg = ((chroma > 10) | ((tex > 7) & (chroma > 4))).astype(np.uint8) * 255
    fg = cv2.morphologyEx(fg, cv2.MORPH_OPEN, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5)))
    fg = cv2.morphologyEx(fg, cv2.MORPH_CLOSE, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (35, 35)))
    H, W = fg.shape
    n, ccl, st, _ = cv2.connectedComponentsWithStats(fg, 8)
    filled = np.zeros((H, W), np.uint8)
    for i in range(1, n):
        if st[i, cv2.CC_STAT_AREA] * MM2_PX < 20:
            continue
        comp = (ccl == i).astype(np.uint8); ff = comp.copy() * 255
        cv2.floodFill(ff, np.zeros((H + 2, W + 2), np.uint8), (0, 0), 255)
        filled |= cv2.bitwise_or(comp * 255, cv2.bitwise_not(ff))
    return filled, a, b


def spot_segment(a, b, cmask):
    """Isolate individual purple guide spots inside one corolla mask."""
    P = (a - b).astype(np.float32); P2 = P.copy(); P2[~cmask] = 0
    se = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (21, 21))
    tophat = P2 - cv2.morphologyEx(P2, cv2.MORPH_OPEN, se)
    inside = P[cmask]; med = np.median(inside); mad = np.median(np.abs(inside - med)) + 1e-6
    spots = (((P > 6) | ((tophat > 2.0) & (P > 2.0)) |
              ((P > med + 3.5 * 1.4826 * mad) & (P > 1.5))) & cmask).astype(np.uint8)
    return cv2.morphologyEx(spots, cv2.MORPH_OPEN, np.ones((2, 2), np.uint8))


def orient_base_tip(cmask, spots):
    """Rotate corolla so its long axis is vertical with the guide-rich base at bottom."""
    ys, xs = np.where(cmask); pts = np.stack([xs, ys], 1).astype(np.float32); mean = pts.mean(0)
    _, _, vt = np.linalg.svd(pts - mean, full_matrices=False); axis = vt[0]
    ang = math.degrees(math.atan2(axis[1], axis[0]))
    M = cv2.getRotationMatrix2D((float(mean[0]), float(mean[1])), ang - 90, 1.0)
    H, W = cmask.shape
    rc = cv2.warpAffine(cmask.astype(np.uint8), M, (W, H), flags=cv2.INTER_NEAREST)
    rs = cv2.warpAffine(spots.astype(np.uint8), M, (W, H), flags=cv2.INTER_NEAREST)
    ys2, xs2 = np.where(rc > 0)
    rc = rc[ys2.min():ys2.max() + 1, xs2.min():xs2.max() + 1]
    rs = rs[ys2.min():ys2.max() + 1, xs2.min():xs2.max() + 1]
    sp_ys = np.where(rs > 0)[0]; hh = rc.shape[0]
    if len(sp_ys) >= 10 and np.mean(sp_ys < hh * 0.4) > np.mean(sp_ys > hh * 0.6):
        rc = rc[::-1]; rs = rs[::-1]
    return rc, rs


def geometry(rc):
    """Provisional geometry from the oriented mask (base at bottom).
    length/width are robust; tube/lobe are provisional (see module docstring)."""
    H, W = rc.shape
    widths = rc.sum(1).astype(float)
    cols_top = np.full(W, -1)
    for x in range(W):
        ys = np.where(rc[:, x] > 0)[0]
        if len(ys):
            cols_top[x] = ys.min()
    valid = cols_top >= 0; prof = cols_top[valid].astype(float)
    if len(prof) > 11:
        kk = max(3, (len(prof) // 25) | 1); prof = cv2.GaussianBlur(prof.reshape(1, -1), (kk, 1), 0).ravel()
    span = prof.max() - prof.min() + 1e-6
    tips = [i for i in range(1, len(prof) - 1) if prof[i] <= prof[i - 1] and prof[i] < prof[i + 1]
            and (prof.max() - prof[i]) > 0.25 * span]
    sinus = [i for i in range(1, len(prof) - 1) if prof[i] >= prof[i - 1] and prof[i] > prof[i + 1]
             and (prof[i] - prof.min()) > 0.25 * span]
    n_lobes = max(len(tips), 1)
    sin_y = np.mean([prof[i] for i in sinus]) if sinus else prof.min() + 0.5 * span
    sr = min(max(int(round(sin_y)), 0), H - 1)
    return dict(length=H, width=widths.max(), throat_width=rc[sr, :].sum(),
                tube_depth=H - 1 - sin_y, n_lobes=n_lobes)


def detect_styles(img, corolla_mask):
    """Thin elongated organ sticks (likely styles/pistils) beside the corollas."""
    Lc, a, b = channels(img); chroma = np.sqrt(a * a + b * b)
    mask = ((chroma > 8) & (corolla_mask == 0)).astype(np.uint8) * 255
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, np.ones((2, 2), np.uint8))
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (9, 3)))
    n, ccl, st, ct = cv2.connectedComponentsWithStats(mask, 8); out = []
    for i in range(1, n):
        amm = st[i, cv2.CC_STAT_AREA] * MM2_PX
        if amm < 1.5 or amm > 60:
            continue
        comp = (ccl == i).astype(np.uint8); cnts, _ = cv2.findContours(comp, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        (rw, rh) = cv2.minAreaRect(max(cnts, key=cv2.contourArea))[1]
        if min(rw, rh) < 1:
            continue
        aspect = max(rw, rh) / min(rw, rh); thick = min(rw, rh) * MM_PX; length_mm = max(rw, rh) * MM_PX
        if aspect < 4.0 or thick > 2.5 or length_mm < 2.0:
            continue
        out.append(dict(cx=float(ct[i][0]), cy=float(ct[i][1]),
                        length_mm=round(length_mm, 2), width_mm=round(thick, 2), aspect=round(aspect, 1)))
    return out


def site_numbers(fname):
    """Site number(s) from a filename. The trailing number(s) = SITE id(s) (地点);
    a range (e.g. oshima10~13) means several sites on one sheet -> assign per corolla
    from the handwritten circled plant numbers in QC. Returns (list[int], is_range, variant)."""
    stem = unicodedata.normalize('NFKC', os.path.splitext(fname)[0]).replace('～', '~').replace('－', '-')
    m = re.search(r'\(([0-9]+)\)', stem); variant = m.group(1) if m else ''
    stem_np = re.sub(r'\([0-9]+\)', '', stem)
    nums = [int(x) for x in re.findall(r'\d+', stem_np)]
    is_range = len(nums) >= 2 and ('~' in stem_np or '-' in stem_np)
    if is_range:
        nums = list(range(min(nums), max(nums) + 1))
    return nums, is_range, variant


# location.xlsx uses short island names; map them to the data-folder names.
LOC_ALIAS = {"shikine": "shikinejima", "kozu": "kozushima"}


def load_locations(path):
    """Read a site-coordinate table (columns: island, no, lat, lon) into {(folder, no): (lat, lon)}.
    'no' is the site number; island names are normalised to the data-folder names."""
    if not path or not os.path.exists(path):
        return {}
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[warn] openpyxl not installed; coordinates skipped"); return {}
    wb = load_workbook(path, data_only=True); ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip().lower() if c is not None else '' for c in rows[0]]
    ci = {name: hdr.index(name) for name in ('island', 'no', 'lat', 'lon') if name in hdr}
    loc = {}; cur = None
    for r in rows[1:]:
        isl = r[ci['island']] if r[ci['island']] not in (None, '') else cur
        cur = isl
        no = r[ci['no']]
        if isl is None or no is None:
            continue
        key = str(isl).strip().lower(); key = LOC_ALIAS.get(key, key)
        try:
            loc[(key, int(no))] = (r[ci['lat']], r[ci['lon']])
        except (TypeError, ValueError):
            continue
    return loc


def process_sheet(path, folder, out_dir, loc_map=None):
    island, order = ISLANDS.get(folder, (folder, ''))
    fname = os.path.basename(path)
    img = load_bgr(path); H, W = img.shape[:2]
    loc_map = loc_map or {}
    snums, is_range, pvar = site_numbers(fname)
    # Islands with a single sampling site: the whole island is one site regardless of
    # the filename range (e.g. Shikinejima = 1 site, 5 individuals). Detected from the
    # coordinate table having exactly one site for that island.
    isl_sites = sorted({n for (isl, n) in loc_map if isl == folder}) if loc_map else []
    if len(isl_sites) == 1:
        snums, is_range = [isl_sites[0]], False
    # site + coords: unambiguous only when the sheet holds a single site
    if len(snums) == 1:
        site_no = snums[0]
        lat, lon = loc_map.get((folder, site_no), ('', ''))
        site_cands = ''
    else:
        site_no = ''; lat = lon = ''
        site_cands = '|'.join(str(n) for n in snums)
    filled, a, b = foreground(img)
    brownmap = (a > 6) & ((a - b) < -15)
    styles = detect_styles(img, filled)
    n, ccl, st, ct = cv2.connectedComponentsWithStats(filled, 8)
    ov = img.copy(); rows = []; k = 0; cents = []
    for i in sorted(range(1, n), key=lambda j: (ct[j][1] // 200, ct[j][0])):
        amm = st[i, cv2.CC_STAT_AREA] * MM2_PX
        if not (AREA_MM2_MIN <= amm <= AREA_MM2_MAX):
            continue
        comp = (ccl == i); cm = comp.astype(np.uint8)
        c0 = max(cv2.findContours(cm, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)[0], key=cv2.contourArea)
        (rw, rh) = cv2.minAreaRect(c0)[1]; aspect = max(rw, rh) / max(1e-6, min(rw, rh))
        ha = cv2.contourArea(cv2.convexHull(c0)); sol = st[i, cv2.CC_STAT_AREA] / ha if ha > 0 else 0
        if aspect > ASPECT_MAX or sol < SOLIDITY_MIN:
            continue
        k += 1; cents.append((ct[i][0], ct[i][1], k))
        spots = spot_segment(a, b, comp)
        sp_area = int(spots.sum()); area_px = int(st[i, cv2.CC_STAT_AREA]); cov = sp_area / area_px
        ns, _, sst, _ = cv2.connectedComponentsWithStats(spots, 8)
        nspot = sum(1 for s in range(1, ns) if sst[s, cv2.CC_STAT_AREA] * MM2_PX >= 0.02)
        brown_frac = int((brownmap & comp).sum()) / area_px
        rc, rs = orient_base_tip(comp, spots.astype(bool))
        g = geometry(rc)
        if rs.sum() > 10:
            sp_ys = np.where(rs > 0)[0]; hh = rc.shape[0]; extent = (hh - 1 - sp_ys.min()) / (hh - 1)
        else:
            extent = ''
        open_flag = g['n_lobes'] >= 4
        circ = g['throat_width'] if open_flag else g['throat_width'] * 2
        len_mm = round(g['length'] * MM_PX, 2); wid_mm = round(g['width'] * MM_PX, 2)
        wl = wid_mm / len_mm if len_mm else 0
        # merged-blob flag: two touching corollas read as one (over-long or wider-than-tall)
        merge_check = 'check' if (len_mm > 55 or wl > 0.95 or amm > 1500) else ''
        rows.append(dict(
            island=island, region_order=order, sheet=os.path.splitext(fname)[0],
            site_no=site_no, site_candidates=site_cands, site_lat=lat, site_lon=lon,
            plant_id='', corolla_id=k, cx=int(ct[i][0]), cy=int(ct[i][1]),
            corolla_len_mm=len_mm,
            corolla_width_mm=wid_mm,
            wl_ratio=round(wl, 3), fold_check=('check' if wl < 0.55 else ''),
            merge_check=merge_check,
            corolla_area_mm2=round(amm, 1),
            guide_area_mm2=round(sp_area * MM2_PX, 2),
            guide_cov_pct=round(cov * 100, 2),
            n_spots=nspot, spot_density_cm2=round(nspot / (amm / 100), 2),
            guide_extent_rel=round(extent, 3) if extent != '' else '',
            guide_present=int(cov * 100 >= 0.5),
            brown_frac=round(brown_frac, 3), degraded_flag=int(brown_frac > 0.10),
            prov_mouth_diam_mm=round(circ / math.pi * MM_PX, 2),
            prov_tube_depth_mm=round(g['tube_depth'] * MM_PX, 2),
            prov_n_lobes=g['n_lobes'],
            solidity=round(sol, 3), aspect=round(aspect, 2), exclude=''))
        cv2.drawContours(ov, [c0], -1, (0, 255, 0), 3)
        cv2.drawContours(ov, cv2.findContours(spots, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)[0], -1, (255, 255, 0), 1)
        cv2.putText(ov, f"{k}", (int(ct[i][0]) - 12, int(ct[i][1])), cv2.FONT_HERSHEY_SIMPLEX, 1.3, (0, 0, 255), 4)
        cv2.putText(ov, f"L{rows[-1]['corolla_len_mm']:.0f} cov{cov*100:.0f} b{brown_frac*100:.0f}",
                    (st[i, cv2.CC_STAT_LEFT], max(18, st[i, cv2.CC_STAT_TOP] - 6)),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.75, (255, 128, 0), 2)
    for s in styles:
        cv2.circle(ov, (int(s['cx']), int(s['cy'])), 6, (0, 140, 255), -1)
        if cents:
            s['nearest_corolla'] = min(cents, key=lambda c: (c[0] - s['cx']) ** 2 + (c[1] - s['cy']) ** 2)[2]
        s['island'] = island; s['sheet'] = os.path.splitext(fname)[0]
        cv2.putText(ov, f"S{s['length_mm']:.0f}", (int(s['cx']) + 6, int(s['cy'])), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 140, 255), 2)
    sc = 1700 / max(H, W)
    ovdir = os.path.join(out_dir, "overlays"); os.makedirs(ovdir, exist_ok=True)
    cv2.imencode('.png', cv2.resize(ov, (int(W * sc), int(H * sc))))[1].tofile(
        os.path.join(ovdir, f"{island}_{os.path.splitext(fname)[0]}.png"))
    return rows, styles


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--data-root", required=True, help="folder containing island sub-folders of scans")
    ap.add_argument("--out-dir", default="results", help="output folder (CSVs + overlays)")
    ap.add_argument("--locations", default="", help="optional site-coordinate xlsx (cols: island,no,lat,lon)")
    args = ap.parse_args()
    os.makedirs(args.out_dir, exist_ok=True)
    loc_map = load_locations(args.locations)
    allrows, allsticks = [], []
    for folder in sorted(os.listdir(args.data_root)):
        d = os.path.join(args.data_root, folder)
        if not os.path.isdir(d):
            continue
        for f in sorted(os.listdir(d)):
            if f.lower().endswith((".jpg", ".jpeg", ".png")):
                r, s = process_sheet(os.path.join(d, f), folder.lower(), args.out_dir, loc_map)
                allrows += r; allsticks += s
    with open(os.path.join(args.out_dir, "traits.csv"), "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=list(allrows[0].keys())); w.writeheader(); [w.writerow(x) for x in allrows]
    with open(os.path.join(args.out_dir, "styles.csv"), "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=['island', 'sheet', 'nearest_corolla', 'cx', 'cy', 'length_mm', 'width_mm', 'aspect'],
                           extrasaction='ignore'); w.writeheader(); [w.writerow(x) for x in allsticks]
    # QC template: one row per corolla. User fills site (for range sheets), individual
    # (circled number), flower no, and fold state, against results/overlays/.
    with open(os.path.join(args.out_dir, "qc_plant_labels.csv"), "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(["island", "sheet", "corolla_id", "cx", "cy",
                    "site_no_auto", "site_candidates", "site_no_FILL", "individual_FILL",
                    "flower_no_FILL", "fold_state_FILL(open/folded)", "split_or_exclude_FILL",
                    "site_lat", "site_lon", "fold_check", "merge_check", "notes"])
        for r in allrows:
            w.writerow([r['island'], r['sheet'], r['corolla_id'], r['cx'], r['cy'],
                        r['site_no'], r['site_candidates'], "", "", "",
                        ("folded" if r['fold_check'] else ""), ("SPLIT?" if r['merge_check'] else ""),
                        r['site_lat'], r['site_lon'], r['fold_check'], r['merge_check'], ""])
    # per-island summary
    import statistics as stt
    isl = {}
    for r in allrows:
        isl.setdefault((r['region_order'], r['island']), []).append(r)
    with open(os.path.join(args.out_dir, "per_island_summary.csv"), "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh); w.writerow(["region_order", "island", "n_corolla", "corolla_len_mm_mean",
                                        "corolla_area_mm2_mean", "guide_cov_pct_mean", "guide_present_frac", "degraded_frac"])
        for (order, island), v in sorted(isl.items()):
            w.writerow([order, island, len(v),
                        round(stt.mean([x['corolla_len_mm'] for x in v]), 1),
                        round(stt.mean([x['corolla_area_mm2'] for x in v]), 0),
                        round(stt.mean([x['guide_cov_pct'] for x in v]), 2),
                        round(stt.mean([x['guide_present'] for x in v]), 2),
                        round(stt.mean([x['degraded_flag'] for x in v]), 2)])
    print(f"corollas={len(allrows)}  sticks={len(allsticks)}  -> {args.out_dir}/")


if __name__ == "__main__":
    main()
